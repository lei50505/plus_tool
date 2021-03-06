#! /usr/bin/env python
# -*- coding: utf-8 -*-

#pylint: disable=broad-except, too-many-instance-attributes, too-many-branches, too-many-statements, too-many-locals

'''doc'''

import os
import time
import traceback

from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers, PatternFill, Side, Border


def to_float(val):
    '''doc'''
    try:
        return float(val)
    except Exception:
        return None

def to_str(val):
    '''doc'''
    if val is None:
        return None
    try:
        val = str(val)
        if val.strip() == "":
            return None
        return val.strip()
    except Exception:
        return None

class Cell():
    '''doc'''
    def __init__(self, cell):
        self.cell = cell

    def get_val(self):
        '''doc'''
        return self.cell.value

    def get_float_val(self):
        '''doc'''
        val = self.get_val()
        return to_float(val)

    def get_str_val(self):
        '''doc'''
        val = self.get_val()
        return to_str(val)

    def set_val(self, val):
        '''doc'''
        self.cell.value = val

    def set_number_format_text(self):
        '''doc'''
        self.cell.number_format = numbers.FORMAT_TEXT

    def set_fill_red(self):
        '''doc'''
        self.cell.fill = PatternFill(fill_type="solid", \
                            start_color="FFCCFF", end_color="FFCCFF")

    def set_fill_blue(self):
        '''doc'''
        self.cell.fill = PatternFill(fill_type="solid", \
                            start_color="CCFFFF", end_color="CCFFFF")

    def set_fill_yellow(self):
        '''doc'''
        self.cell.fill = PatternFill(fill_type="solid", \
                            start_color="FFFFCC", end_color="FFFFCC")

    def set_border_thin(self):
        '''doc'''
        thin_side = Side(border_style="thin", color="000000")
        thin_border = Border(top=thin_side, left=thin_side, \
                            right=thin_side, bottom=thin_side)
        self.cell.border = thin_border

class Sheet():
    '''doc'''
    def __init__(self, sheet):
        self.sheet = sheet

        self.num_col_index = None

        self.num_val_dict = None
        self.num_val_set = None
        self.num_val_list = None

        self.diff_num_rows = None

        self.num_row_dict = None
        self.num_row_list = None

        self.num_row_data = None

        self.copy_row_count = 0

    def cell(self, row, col):
        '''doc'''
        sheet_cell = self.sheet.cell(row=row, column=col)
        cell = Cell(sheet_cell)
        return cell

    def get_max_col(self):
        '''doc'''
        return self.sheet.max_column

    def get_max_row(self):
        '''doc'''
        return self.sheet.max_row

    def init_num_col_index(self):
        '''doc'''
        max_col = self.get_max_col()
        max_row = self.get_max_row()

        num_col_count = 0
        num_col_index = 0

        for col_index in range(1, max_col + 1):
            num_cell_count = 0
            is_str_cell = False
            for row_index in range(1, max_row + 1):
                cell = self.cell(row_index, col_index)
                float_val = cell.get_float_val()

                if isinstance(float_val, float):
                    num_cell_count += 1
                    continue

                str_val = cell.get_str_val()

                if isinstance(str_val, str):
                    is_str_cell = True
                    break



            if num_cell_count >= 1 and not is_str_cell:
                num_col_count += 1
                num_col_index = col_index

        if num_col_count == 1:
            self.num_col_index = num_col_index

        if num_col_count == 0:
            raise Exception("没有数字列")

        if num_col_count > 1:
            raise Exception("有%d列是数字" % num_col_count)



    def init_num_col_dict(self):
        '''doc'''

        if self.num_col_index is None:
            raise Exception("请先初始化num_col_index")

        self.num_val_dict = {}
        self.num_val_set = set()
        self.num_val_list = []

        self.num_row_dict = {}
        self.num_row_list = []

        self.num_row_data = []

        num_col_index = self.num_col_index
        max_row = self.get_max_row()
        for row_index in range(1, max_row + 1):
            cell = self.cell(row_index, num_col_index)
            float_val = cell.get_float_val()
            if isinstance(float_val, float):
                self.num_val_set.add(float_val)
                self.num_val_list.append(float_val)
                self.num_row_dict[row_index] = float_val
                self.num_row_list.append(row_index)

                tmp_data = {}
                tmp_data["val"] = float_val
                tmp_data["index"] = row_index
                self.num_row_data.append(tmp_data)
                dict_val = self.num_val_dict.get(float_val)
                if dict_val is None:
                    self.num_val_dict[float_val] = 1
                    continue

                self.num_val_dict[float_val] = dict_val + 1


    def init_diff_num_rows(self):
        '''doc'''
        self.diff_num_rows = []
        for num_row in self.num_row_list:
            val = self.num_row_dict.get(num_row)
            count = self.num_val_list.count(val)
            if count == 1:
                self.diff_num_rows.append(num_row)

    def get_row_list_by_val(self, val):
        '''doc'''
        val = to_float(val)
        ret = []
        for num_row in self.num_row_list:
            vas = self.num_row_dict.get(num_row)
            if vas == val:
                ret.append(num_row)
        return ret

    def copy_row_from_sheet(self, src_sheet, row_index, color):
        '''doc'''

        copy_row_count = self.copy_row_count + 1
        src_sheet_max_col = src_sheet.get_max_col()

        for src_sheet_col_index in range(1, src_sheet_max_col + 1):
            src_cell = src_sheet.cell(row_index, src_sheet_col_index)
            tar_cell = self.cell(copy_row_count, src_sheet_col_index)
            tar_cell.set_val(src_cell.get_val())

            tar_cell.set_border_thin()

            tar_cell.set_number_format_text()

            if color == "red":
                tar_cell.set_fill_red()
            elif color == "blue":
                tar_cell.set_fill_blue()
            elif color == "yellow":
                tar_cell.set_fill_yellow()

        self.copy_row_count += 1



class Book():
    '''doc'''
    def __init__(self, book):
        self.book = book

    def active(self):
        '''doc'''
        active_sheet = self.book.active
        return Sheet(active_sheet)

    def sheet(self, sheet_name):
        '''doc'''

        book_sheet = self.book[sheet_name]
        sheet = Sheet(book_sheet)
        return sheet


    def has_sheet(self, *sheet_names):
        '''doc'''
        sheet_names_len = len(sheet_names)
        if sheet_names_len == 0:
            return True

        book_sheet_names = self.book.get_sheet_names()
        for sheet_name in sheet_names:
            if not isinstance(sheet_name, str):
                return False
            if sheet_name not in book_sheet_names:
                return False
        return True

    def save(self, path):
        '''doc'''
        self.book.save(path)

    def close(self):
        '''doc'''
        if self.book is not None:
            self.book.close()

def create_book():
    '''doc'''
    work_book = Workbook(write_only=False)
    book = Book(work_book)
    return book

def load_book(file_path):
    '''doc'''
    work_book = load_workbook(file_path, read_only=True, keep_vba=False, \
                    data_only=True, guess_types=False, keep_links=False)
    book = Book(work_book)
    return book

def get_by_sum(src_arr, tar_sum):
    '''doc'''
    src_arr = sorted(src_arr, key=lambda item: item["key"])
    final_result = []
    cur_result = []
    def iter_by_sum(cur_sum, cur_index):
        '''doc'''
        if cur_index >= len(src_arr):
            return
        cur_val = src_arr[cur_index]["key"]
        if cur_sum < cur_val:
            return
        if cur_sum == cur_val:
            tmp = []
            tmp.append(src_arr[cur_index]["data"])

            for cur_result_item in cur_result:
                tmp.append(src_arr[cur_result_item]["data"])
            final_result.append(tmp)

        cur_result.append(cur_index)
        iter_by_sum(cur_sum - cur_val, cur_index + 1)
        cur_result.pop()
        iter_by_sum(cur_sum, cur_index + 1)

    iter_by_sum(tar_sum, 0)
    return final_result


def main():
    '''doc'''
    if not os.path.isfile("in.xlsx"):
        print("in.xlsx不存在")
        return

    in_book = load_book("in.xlsx")

    if not in_book.has_sheet("Sheet1"):
        print("in.xlsx中不存在Sheet1")
        return

    if not in_book.has_sheet("Sheet2"):
        print("in.xlsx中不存在Sheet2")
        return

    in_sheet1 = in_book.sheet("Sheet1")

    print("初始化Sheet1数字列序号")
    in_sheet1.init_num_col_index()

    in_sheet2 = in_book.sheet("Sheet2")

    print("初始化Sheet2数字列序号")
    in_sheet2.init_num_col_index()


    out_book = create_book()

    out_sheet = out_book.active()

    sheet1_row_done = {}
    sheet2_row_done = {}

    sheet1_max_row = in_sheet1.get_max_row()
    sheet2_max_row = in_sheet2.get_max_row()

    for row_index in range(1, sheet1_max_row + 1):
        sheet1_row_done[row_index] = False

    for row_index in range(1, sheet2_max_row + 1):
        sheet2_row_done[row_index] = False

    print("初始化Sheet1数字列数据")
    in_sheet1.init_num_col_dict()

    print("初始化Sheet2数字列数据")
    in_sheet2.init_num_col_dict()

    print("正在处理数据")

    src_arr = []
    for num_row in in_sheet1.num_row_data:
        src_arr_item = {}
        src_arr_item["key"] = num_row["val"]
        src_arr_item["data"] = num_row["index"]
        src_arr.append(src_arr_item)



    for num_row in in_sheet2.num_row_data:
        tar_arr = get_by_sum(src_arr, num_row["val"])
        color_index = 0
        out_sheet.copy_row_from_sheet(in_sheet2, num_row["index"], "red")
        sheet2_row_done[num_row["index"]] = True
        for tar_item in tar_arr:
            color_index += 1
            for tar_index in tar_item:
                sheet1_row_done[tar_index] = True
                if color_index % 2 == 0:
                    out_sheet.copy_row_from_sheet(in_sheet1, tar_index, "yellow")
                else:
                    out_sheet.copy_row_from_sheet(in_sheet1, tar_index, "blue")

    for row_index in range(1, sheet1_max_row):
        if sheet1_row_done[row_index]:

            continue


        sheet1_row_done[row_index] = True
        out_sheet.copy_row_from_sheet(in_sheet1, row_index, "white")
    for row_index in range(1, sheet2_max_row):
        if sheet2_row_done[row_index]:
            continue


        sheet2_row_done[row_index] = True
        out_sheet.copy_row_from_sheet(in_sheet2, row_index, "white")

    out_book.save("out.xlsx")
    out_book.close()
    in_book.close()


if __name__ == "__main__":
    try:
        main()
    except Exception:
        print(traceback.format_exc())
        time.sleep(120)
    print("OK")
    time.sleep(2)
